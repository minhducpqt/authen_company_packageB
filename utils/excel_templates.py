import io
import pandas as pd
from datetime import datetime
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import httpx, os

SERVICE_A_BASE_URL = os.getenv("SERVICE_A_BASE_URL", "http://127.0.0.1:8824")


async def _next_project_code(access: str, company_code: str) -> str:
    """
    Trả về project_code dạng COMPANYCODEN (N>=1) nhỏ nhất chưa tồn tại.
    """
    headers = {"Authorization": f"Bearer {access}"}
    async with httpx.AsyncClient(base_url=SERVICE_A_BASE_URL, timeout=10.0) as client:
        r = await client.get(
            "/api/v1/projects",
            params={"company_code": company_code, "size": 1000},
            headers=headers,
        )
        if r.status_code == 200:
            data = r.json()
            existing = {p["project_code"].upper() for p in data.get("data", [])}
        else:
            existing = set()

    base = company_code.upper()
    n = 1
    while True:
        cand = f"{base}{n}"
        if cand not in existing:
            return cand
        n += 1


async def build_projects_lots_template(access: str, company_code: str) -> StreamingResponse:
    """
    Sinh file Excel template đẹp cho import dự án + lô
    - project_code mẫu = COMPANYCODEN chưa tồn tại
    - Dự án mặc định INACTIVE, lô mặc định ACTIVE (sẽ do API xử lý sau)
    """
    proj_code = await _next_project_code(access, company_code)

    # ---- Cấu trúc cột
    proj_cols = ["project_code", "name", "description", "location"]
    lot_cols = [
        "project_code",
        "lot_code",
        "name",
        "description",
        "starting_price",
        "deposit_amount",
        "area",
        "bid_step_vnd",   # ⭐ NEW: bước giá mỗi lô
    ]

    # ---- Dòng mẫu
    dfp = pd.DataFrame(
        [
            {
                "project_code": proj_code,
                "name": "Dự án mẫu",
                "description": "",
                "location": "Quận 1, TP.HCM",
            }
        ],
        columns=proj_cols,
    )

    dfl = pd.DataFrame(
        [
            {
                "project_code": proj_code,
                "lot_code": "L-A01",
                "name": "Lô A01",
                "description": "",
                "starting_price": 1_500_000_000,
                "deposit_amount": 150_000_000,
                "area": 80.0,
                "bid_step_vnd": 50_000_000,  # ⭐ Ví dụ bước giá
            }
        ],
        columns=lot_cols,
    )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        dfp.to_excel(w, index=False, sheet_name="projects")
        dfl.to_excel(w, index=False, sheet_name="lots")

        wb = w.book
        ws_p = w.sheets["projects"]
        ws_l = w.sheets["lots"]

        # ===== Styles =====
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        thin = Side(style="thin", color="D9D9D9")
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        money_style = NamedStyle(name="money_vn")
        money_style.number_format = "#,##0"
        money_style.alignment = Alignment(horizontal="right")
        if "money_vn" not in wb.named_styles:
            wb.add_named_style(money_style)

        area_style = NamedStyle(name="area_vn")
        area_style.number_format = "#,##0.00"
        area_style.alignment = Alignment(horizontal="right")
        if "area_vn" not in wb.named_styles:
            wb.add_named_style(area_style)

        wrap_left = Alignment(wrap_text=True, horizontal="left", vertical="top")

        def format_sheet(ws, widths, money_cols=None, area_cols=None):
            max_col = ws.max_column
            max_row = ws.max_row

            # Header
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Column widths
            for idx, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width

            # Freeze & filter
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

            # Body borders + wrap cho description
            desc_col_idx = None
            for c in range(1, max_col + 1):
                if (ws.cell(row=1, column=c).value or "").strip().lower() == "description":
                    desc_col_idx = c
                    break

            for r in range(2, max_row + 1):
                for c in range(1, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = cell_border
                    if desc_col_idx and c == desc_col_idx:
                        cell.alignment = wrap_left

            # Number formats
            for col in (money_cols or []):
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=col).style = money_style

            for col in (area_cols or []):
                for r in range(2, max_row + 1):
                    ws.cell(row=r, column=col).style = area_style

        # Apply style cho projects
        format_sheet(ws_p, widths=[14, 26, 40, 22])

        # Apply style cho lots
        format_sheet(
            ws_l,
            widths=[14, 14, 20, 36, 16, 16, 12, 16],  # ⭐ thêm width cho cột bid_step_vnd
            money_cols=[5, 6, 8],                     # ⭐ starting_price, deposit_amount, bid_step_vnd
            area_cols=[7],
        )

    buf.seek(0)
    filename = f"auction_import_template_{company_code}_{datetime.now():%Y%m%d}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename=\"{filename}\"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
