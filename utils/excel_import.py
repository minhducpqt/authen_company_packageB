# utils/excel_import.py
"""Import dự án từ Excel — parse + preview verify (Service B). Không đụng logic Service A."""
from __future__ import annotations

import json
import os
import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import httpx
from openpyxl import load_workbook

from utils.project_import_verifier import ProjectImportVerifier, is_strict_non_negative_integer

SERVICE_A_BASE_URL = os.getenv("SERVICE_A_BASE_URL", "http://127.0.0.1:8824")

REQUIRED_PROJECT_HEADERS = ["project_code", "name"]
REQUIRED_LOT_HEADERS = ["project_code", "lot_code", "name", "starting_price", "deposit_amount"]


def _strip_accents(s: str) -> str:
    if s is None:
        return ""
    nkfd = unicodedata.normalize("NFD", str(s))
    no_acc = "".join(ch for ch in nkfd if not unicodedata.combining(ch))
    return unicodedata.normalize("NFC", no_acc)


def normalize_code(code: str) -> str:
    if code is None:
        return ""
    code = _strip_accents(code).upper()
    return code.replace(" ", "")


def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    return " ".join(s.split())


def _headerize(s: str) -> str:
    s = _strip_accents(str(s)).strip().lower()
    for ch in "-./":
        s = s.replace(ch, " ")
    return "_".join(s.split())


async def _get_json(client: httpx.AsyncClient, url: str, headers: Dict[str, str]):
    r = await client.get(url, headers=headers)
    try:
        return r.status_code, r.json()
    except Exception:
        return r.status_code, None


def _parse_money_cell(v: Any) -> Any:
    """
    Giữ số nguyên nguyên; không round 123.5 → 124.
    Returns:
      int | None | 'NaN' (không parse được / không phải số nguyên)
    """
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return "NaN"
    if isinstance(v, int):
        return int(v) if v >= 0 else "NaN"
    if isinstance(v, float):
        if v >= 0 and v.is_integer():
            return int(v)
        return "NaN"
    s = str(v).strip().replace(",", "").replace(" ", "")
    if not s:
        return None
    try:
        f = float(s)
    except Exception:
        return "NaN"
    if f < 0 or not float(f).is_integer():
        # check exact integer string via Decimal path
        if not is_strict_non_negative_integer(s):
            return "NaN"
        try:
            return int(float(s)) if float(s).is_integer() else "NaN"
        except Exception:
            return "NaN"
    return int(f)


def _to_float(v: Any):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace(" ", ""))
    except Exception:
        return "NaN"


def _read_sheets(file_bytes: bytes) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[str]]:
    """projects, lots, template_errors"""
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    except Exception:
        return [], [], ["File tải lên không phải Excel hợp lệ (.xlsx/.xls)."]

    names = [ws.title.strip().lower() for ws in wb.worksheets]
    name_map = {ws.title.strip().lower(): ws.title for ws in wb.worksheets}
    if "projects" not in names or "lots" not in names:
        return [], [], ["Template không hợp lệ. Cần có đủ 2 sheet: 'projects' và 'lots'."]

    ws_p = wb[name_map["projects"]]
    ws_l = wb[name_map["lots"]]

    headers_p_raw = list(next(ws_p.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers_l_raw = list(next(ws_l.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers_p = [_headerize(h) if h is not None else f"col_{i}" for i, h in enumerate(headers_p_raw)]
    headers_l = [_headerize(h) if h is not None else f"col_{i}" for i, h in enumerate(headers_l_raw)]

    errors: List[str] = []
    miss_p = [h for h in REQUIRED_PROJECT_HEADERS if h not in headers_p]
    miss_l = [h for h in REQUIRED_LOT_HEADERS if h not in headers_l]
    if miss_p:
        errors.append(f"Sheet 'projects' thiếu cột bắt buộc: {', '.join(miss_p)}")
    if miss_l:
        errors.append(f"Sheet 'lots' thiếu cột bắt buộc: {', '.join(miss_l)}")
    if errors:
        return [], [], errors

    projects: List[Dict[str, Any]] = []
    for excel_row, row in enumerate(ws_p.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        rec = {headers_p[i]: row[i] for i in range(min(len(headers_p), len(row)))}
        rec["project_code"] = normalize_code(rec.get("project_code", ""))
        rec["name"] = normalize_text(rec.get("name", ""))
        rec["description"] = normalize_text(rec.get("description", ""))
        rec["location"] = normalize_text(rec.get("location", ""))
        rec["row"] = excel_row
        projects.append(rec)

    lots: List[Dict[str, Any]] = []
    for excel_row, row in enumerate(ws_l.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        rec = {headers_l[i]: row[i] for i in range(min(len(headers_l), len(row)))}
        rec["project_code"] = normalize_code(rec.get("project_code", ""))
        rec["lot_code"] = normalize_code(rec.get("lot_code", ""))
        rec["name"] = normalize_text(rec.get("name", ""))
        rec["description"] = normalize_text(rec.get("description", ""))
        rec["starting_price"] = _parse_money_cell(rec.get("starting_price"))
        rec["deposit_amount"] = _parse_money_cell(rec.get("deposit_amount"))
        rec["area"] = _to_float(rec.get("area"))
        rec["bid_step_vnd"] = _parse_money_cell(rec.get("bid_step_vnd")) if rec.get("bid_step_vnd") not in (None, "") else None
        rec["row"] = excel_row
        lots.append(rec)

    return projects, lots, []


def _validate_structure(projects: List[Dict[str, Any]], lots: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Lỗi cấu trúc/text (không thay rules số của ProjectImportVerifier)."""
    errors: List[Dict[str, Any]] = []
    seen = set()
    for p in projects:
        idx = p.get("row") or 0
        if not p.get("project_code"):
            errors.append(
                {
                    "type": "projects",
                    "row": idx,
                    "sheet": "projects",
                    "field": "project_code",
                    "code": "MISSING_PROJECT_CODE",
                    "msg": "Thiếu project_code",
                }
            )
        if not p.get("name"):
            errors.append(
                {
                    "type": "projects",
                    "row": idx,
                    "sheet": "projects",
                    "field": "name",
                    "code": "MISSING_PROJECT_NAME",
                    "msg": "Thiếu name",
                }
            )
        code = p.get("project_code") or ""
        if code:
            if code in seen:
                errors.append(
                    {
                        "type": "projects",
                        "row": idx,
                        "sheet": "projects",
                        "field": "project_code",
                        "code": "DUPLICATE_PROJECT_CODE",
                        "msg": "Trùng project_code trong file",
                    }
                )
            seen.add(code)

    pset = {p.get("project_code") for p in projects if p.get("project_code")}
    for l in lots:
        idx = l.get("row") or 0
        if not l.get("project_code"):
            errors.append(
                {
                    "type": "lots",
                    "row": idx,
                    "sheet": "lots",
                    "field": "project_code",
                    "code": "MISSING_LOT_PROJECT_CODE",
                    "msg": "Thiếu project_code",
                }
            )
        elif l.get("project_code") not in pset:
            errors.append(
                {
                    "type": "lots",
                    "row": idx,
                    "sheet": "lots",
                    "field": "project_code",
                    "code": "UNKNOWN_PROJECT_CODE",
                    "msg": f"project_code '{l.get('project_code')}' không có trong sheet projects",
                }
            )
        if not l.get("lot_code"):
            errors.append(
                {
                    "type": "lots",
                    "row": idx,
                    "sheet": "lots",
                    "field": "lot_code",
                    "code": "MISSING_LOT_CODE",
                    "msg": "Thiếu lot_code",
                }
            )
        if not l.get("name"):
            errors.append(
                {
                    "type": "lots",
                    "row": idx,
                    "sheet": "lots",
                    "field": "name",
                    "code": "MISSING_LOT_NAME",
                    "msg": "Thiếu name",
                }
            )
        bstep = l.get("bid_step_vnd")
        if bstep == "NaN":
            errors.append(
                {
                    "type": "lots",
                    "row": idx,
                    "sheet": "lots",
                    "field": "bid_step_vnd",
                    "code": "INVALID_BID_STEP",
                    "msg": "bid_step_vnd không phải số nguyên hợp lệ",
                }
            )
        elif isinstance(bstep, int) and bstep < 0:
            errors.append(
                {
                    "type": "lots",
                    "row": idx,
                    "sheet": "lots",
                    "field": "bid_step_vnd",
                    "code": "INVALID_BID_STEP",
                    "msg": "bid_step_vnd phải >= 0",
                }
            )
    return errors


async def _get_project_by_code(access: str, code: str) -> Optional[Dict[str, Any]]:
    headers = {"Authorization": f"Bearer {access}"}
    async with httpx.AsyncClient(base_url=SERVICE_A_BASE_URL, timeout=12.0) as client:
        st, data = await _get_json(client, f"/api/v1/projects/by_code/{code}", headers)
        if st == 200 and isinstance(data, dict):
            return data
        return None


def _lots_for_apply(verified_lots: List[Dict[str, Any]], original_lots: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Payload apply: giữ field Service A quen thuộc, dùng int đã verify khi có."""
    by_row = {v.get("row"): v for v in verified_lots}
    out: List[Dict[str, Any]] = []
    for orig in original_lots:
        v = by_row.get(orig.get("row")) or {}
        sp = v.get("starting_price_int")
        dp = v.get("deposit_amount_int")
        if sp is None and isinstance(orig.get("starting_price"), int):
            sp = orig.get("starting_price")
        if dp is None and isinstance(orig.get("deposit_amount"), int):
            dp = orig.get("deposit_amount")
        out.append(
            {
                "project_code": orig.get("project_code"),
                "lot_code": orig.get("lot_code"),
                "name": orig.get("name"),
                "description": orig.get("description"),
                "starting_price": sp,
                "deposit_amount": dp,
                "bid_step_vnd": orig.get("bid_step_vnd") if orig.get("bid_step_vnd") != "NaN" else None,
                "area": orig.get("area") if orig.get("area") != "NaN" else None,
                "row": orig.get("row"),
            }
        )
    return out


async def handle_import_preview(file_bytes: bytes, access: str) -> Dict[str, Any]:
    """
    Parse Excel → structural validate → ProjectImportVerifier → check ACTIVE conflict trên A.
    Luôn trả đủ preview khi template OK (kể cả khi có ERROR lô).
    """
    projects, lots, tpl_errs = _read_sheets(file_bytes)
    if tpl_errs:
        return {
            "ok": False,
            "template_error": True,
            "errors": [{"msg": e} for e in tpl_errs],
            "projects": [],
            "lots": [],
            "conflicts_active": [],
            "conflicts_inactive": [],
            "verification": None,
            "can_continue": False,
        }

    struct_errs = _validate_structure(projects, lots)
    verification = ProjectImportVerifier(lots).run()

    # Flat errors for legacy UI list
    flat_errors: List[Dict[str, Any]] = list(struct_errs)
    for lr in verification.get("lots") or []:
        for e in lr.get("errors") or []:
            flat_errors.append(
                {
                    "type": "lots",
                    "sheet": "lots",
                    "row": lr.get("row"),
                    "field": e.get("field"),
                    "code": e.get("code"),
                    "msg": e.get("message") or e.get("msg"),
                    "lot_code": lr.get("lot_code"),
                    "project_code": lr.get("project_code"),
                }
            )

    conflicts_active: List[str] = []
    conflicts_inactive: List[str] = []
    codes = sorted({p["project_code"] for p in projects if p.get("project_code")})
    for code in codes:
        ex = await _get_project_by_code(access, code)
        if not ex:
            continue
        status = (ex.get("status") or "").upper()
        if status == "ACTIVE":
            conflicts_active.append(code)
        else:
            conflicts_inactive.append(code)

    has_struct = len(struct_errs) > 0
    has_verify_err = bool(verification.get("has_errors"))
    has_block = len(conflicts_active) > 0
    can_continue = (
        (not has_struct)
        and (not has_verify_err)
        and (not has_block)
        and bool(verification.get("can_continue", True))
    )

    apply_lots = _lots_for_apply(verification.get("lots") or [], lots)

    return {
        "ok": can_continue,
        "template_error": False,
        "errors": flat_errors,
        "projects": projects,
        "lots": apply_lots,
        "lot_preview": verification.get("lots") or [],
        "conflicts_active": conflicts_active,
        "conflicts_inactive": conflicts_inactive,
        "verification": verification,
        "can_continue": can_continue,
        "summary": {
            "totalLots": verification.get("totalLots"),
            "errorCount": verification.get("errorCount") + len(struct_errs),
            "warningCount": verification.get("warningCount"),
            "totalStartingPrice": verification.get("totalStartingPrice"),
            "totalDeposit": verification.get("totalDeposit"),
            "projectDepositPercent": verification.get("projectDepositPercent"),
            "representativeDepositPercent": verification.get("representativeDepositPercent"),
            "projectWarnings": verification.get("projectWarnings") or [],
        },
    }


# Gợi ý payload_json cho form apply (tránh nhét full verification khổng lồ nếu không cần)
def preview_payload_for_apply(preview: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "ok": preview.get("ok"),
        "projects": preview.get("projects") or [],
        "lots": preview.get("lots") or [],
        "conflicts_active": preview.get("conflicts_active") or [],
        "conflicts_inactive": preview.get("conflicts_inactive") or [],
        "errors": preview.get("errors") or [],
    }


def dumps_preview_payload(preview: Dict[str, Any]) -> str:
    return json.dumps(preview_payload_for_apply(preview), ensure_ascii=False)
